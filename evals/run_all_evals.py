"""
run_all_evals.py

Transcribes clean and distorted once, evaluates every restored_* directory,
writes a consolidated Excel workbook, and prints a summary table.

Usage
-----
    python3 evals/run_all_evals.py
    python3 evals/run_all_evals.py --output results/
"""

from __future__ import annotations

import argparse
import csv
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import eval_runner as er

DATA_ROOT = Path("data")

CSV_FIELDNAMES = ["condition", "file", "action_id", "ground_truth",
                  "predicted", "correct", "wer", "transcript"]

# Excel row fill colours keyed by condition name
_FILL_HEX = {"clean": "D9EAD3", "distorted": "FCE5CD"}
_FILL_HEX_DEFAULT = "CFE2F3"


def discover_restored_dirs() -> list[tuple[str, Path]]:
    """Return [(label, path)] for each restored_* dir containing WAVs.

    data/restored/ is skipped when its WAV set is identical to restored_auto_any
    (it is populated from that source by default and would be a duplicate run).
    """
    default_dir = DATA_ROOT / "restored"
    auto_any_dir = DATA_ROOT / "restored_auto_any"

    default_wavs = {f.name for f in default_dir.glob("*.wav")} if default_dir.exists() else set()
    auto_any_wavs = {f.name for f in auto_any_dir.glob("*.wav")} if auto_any_dir.exists() else set()

    dirs: list[tuple[str, Path]] = []
    if default_wavs and default_wavs != auto_any_wavs:
        dirs.append(("restored", default_dir))

    for d in sorted(DATA_ROOT.glob("restored_*")):
        if d.is_dir() and any(d.glob("*.wav")):
            dirs.append((d.name[len("restored_"):], d))

    return dirs


def _make_summary_row(label: str, results: list[dict]) -> dict:
    n = len(results)
    correct = sum(r["correct"] for r in results)
    wer_vals = [r["wer"] for r in results if r.get("wer") is not None]
    return {
        "condition": label,
        "n": n,
        "correct": correct,
        "car": round(correct / n, 4) if n else 0.0,
        "mean_wer": round(sum(wer_vals) / len(wer_vals), 4) if wer_vals else None,
    }


def print_summary_table(rows: list[dict]) -> None:
    col_w = {"condition": 22, "n": 5, "correct": 8, "car": 8, "wer": 9}
    header = (
        f"{'Condition':<{col_w['condition']}}"
        f"{'N':>{col_w['n']}}"
        f"{'Correct':>{col_w['correct']}}"
        f"{'CAR':>{col_w['car']}}"
        f"{'Mean WER':>{col_w['wer']}}"
    )
    sep = "-" * len(header)
    print("\n" + sep)
    print(header)
    print(sep)
    for r in rows:
        wer_str = f"{r['mean_wer']:.3f}" if r["mean_wer"] is not None else "  n/a"
        print(
            f"{r['condition']:<{col_w['condition']}}"
            f"{r['n']:>{col_w['n']}}"
            f"{r['correct']:>{col_w['correct']}}"
            f"{r['car']:>{col_w['car']}.1%}"
            f"{wer_str:>{col_w['wer']}}"
        )
    print(sep + "\n")


def _style_header_row(ws, n_cols: int, hdr_font, hdr_fill) -> None:
    from openpyxl.styles import Alignment
    for col in range(1, n_cols + 1):
        cell = ws.cell(1, col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")


def write_excel(summary_rows: list[dict], all_raw: list[dict], output_path: Path) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F5597")

    wb = openpyxl.Workbook()

    ws_sum = wb.active
    ws_sum.title = "Summary"
    sum_headers = ["Condition", "N", "Correct", "CAR", "Mean WER"]
    ws_sum.append(sum_headers)
    _style_header_row(ws_sum, len(sum_headers), hdr_font, hdr_fill)

    for r in summary_rows:
        wer_val = r["mean_wer"] if r["mean_wer"] is not None else ""
        ws_sum.append([r["condition"], r["n"], r["correct"], r["car"], wer_val])
        row_idx = ws_sum.max_row
        fill_hex = _FILL_HEX.get(r["condition"], _FILL_HEX_DEFAULT)
        row_fill = PatternFill("solid", fgColor=fill_hex)
        for col in range(1, 6):
            ws_sum.cell(row_idx, col).fill = row_fill
        ws_sum.cell(row_idx, 4).number_format = "0.0%"
        if wer_val:
            ws_sum.cell(row_idx, 5).number_format = "0.000"

    for col, width in zip(range(1, 6), [24, 6, 9, 9, 10]):
        ws_sum.column_dimensions[get_column_letter(col)].width = width

    ws_raw = wb.create_sheet("Raw")
    ws_raw.append(CSV_FIELDNAMES)
    _style_header_row(ws_raw, len(CSV_FIELDNAMES), hdr_font, hdr_fill)
    for row in all_raw:
        ws_raw.append([row.get(h, "") for h in CSV_FIELDNAMES])
    for col_cells in ws_raw.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws_raw.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 60)

    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Run all eval conditions and write consolidated Excel results."
    )
    parser.add_argument("--output", default="results/",
                        help="Output directory (default: results/).")
    args = parser.parse_args()

    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    restored_dirs = discover_restored_dirs()
    if not restored_dirs:
        print("[ERROR] No restored_* directories with WAV files found under data/")
        sys.exit(1)

    print(f"Found {len(restored_dirs)} restored condition(s):")
    for label, path in restored_dirs:
        print(f"  {label:25s}  {path}")

    all_raw: list[dict] = []
    summary_rows: list[dict] = []
    reference_transcripts: dict[str, str] = {}

    for cond in ("clean", "distorted"):
        data_dir = er.DATA_DIRS[cond]
        if not data_dir.exists():
            print(f"\n[SKIP] {cond}: {data_dir} not found")
            continue
        print(f"\nRunning condition: {cond} ({data_dir})")
        results = er.run_condition(
            cond, data_dir,
            reference_transcripts=reference_transcripts if cond != "clean" else None,
        )
        if cond == "clean":
            reference_transcripts = {Path(r["file"]).stem: r["transcript"] for r in results}
            print(f"  {len(reference_transcripts)} clean references ready for WER.")
        all_raw.extend(results)
        summary_rows.append(_make_summary_row(cond, results))
        print(f"  {len(results)} samples evaluated")

    print(f"\nRunning {len(restored_dirs)} restored condition(s) …")
    for label, path in restored_dirs:
        print(f"\n  Condition: {label} ({path})")
        results = er.run_condition(label, path, reference_transcripts=reference_transcripts)
        all_raw.extend(results)
        summary_rows.append(_make_summary_row(label, results))
        print(f"    {len(results)} samples evaluated")

    print_summary_table(summary_rows)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    raw_path = output_dir / f"all_raw_{timestamp}.csv"
    with open(raw_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDNAMES)
        writer.writeheader()
        writer.writerows(all_raw)

    xlsx_path = output_dir / f"all_results_{timestamp}.xlsx"
    write_excel(summary_rows, all_raw, xlsx_path)

    print(f"Raw CSV  -> {raw_path}")
    print(f"Excel    -> {xlsx_path}")


if __name__ == "__main__":
    main()
